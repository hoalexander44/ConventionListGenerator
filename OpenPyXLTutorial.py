### PART 1

#from openpyxl import Workbook, load_workbook

#wb = load_workbook('conventions.xlsx')
#ws = wb.active
#print(ws)
#ws['A2'].value = "test convention name"

#wb.save('conventions.xlsx')


### PART 2

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

current_time = datetime.now()
#date_time_string = current_time.strftime("%d%m%Y%H%M%S")
date_time_string = ""

wb = Workbook()
ws = wb.active
ws.title = "Conventions"

ws.append(["Convention Name", "Date", "Convention Center", "City and State"])

for row in range(2, 11):
	for col in range(1, 5):
		char = get_column_letter(col)
		ws[char + str(row)] = "test"



wb_string = "conventionList_" + date_time_string + ".xlsx"
wb.save(wb_string)
