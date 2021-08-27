from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

import time

if __name__=="__main__":
	PATH = ".\chromedriver.exe"
	driver = webdriver.Chrome(PATH)

	driver.get("https://animecons.com/events/schedule.php?loc=us&year=2021")
	con_list_table_element = driver.find_element_by_id("ConListTable")

	# odd and even doesn't matter its just how the website organized it

	con_list_table_body = con_list_table_element.find_element_by_tag_name('tbody')
	conventions = con_list_table_body.find_elements_by_tag_name("tr")
	#conventions = con_list_table_element.find_elements_by_class_name("even")
	#oddConventions = con_list_table_element.find_elements_by_class_name("odd")

	#conventions.extend(oddConventions)

	processed_convention_list = []

	for convention in conventions:
		contents = convention.find_elements_by_tag_name('td')
		contentList = []
		for content in contents:
			print(content.text)
			contentList.append(content.text)
		processed_convention_list.append(contentList)

		print("\n")

	driver.quit()


	print(processed_convention_list)


	wb = Workbook()
	ws = wb.active
	ws.title = "Conventions"

	ws.append(["Convention Name", "Date", "Location"])

	for row in range(0, len(conventions)):
		for col in range(0, 3):
			char = get_column_letter(col + 1)
			ws[char + str(row +2)] = processed_convention_list[row][col]



	wb_string = "conventionList_.xlsx"
	wb.save(wb_string)

